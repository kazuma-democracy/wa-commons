from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Iterable

from .jpx import from_jpx_row
from .models import EntityRecord, SourceRef
from .snapshots import sha256_file

DOMESTIC_MARKET_MARKER = "内国株式"


def _read_csv(path: Path) -> list[dict[str, object]]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        return list(csv.DictReader(fh))


def _rows_from_matrix(values: list[list[object]]) -> list[dict[str, object]]:
    if not values:
        return []
    headers = [str(v).strip() if v is not None else "" for v in values[0]]
    return [
        {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        for row in values[1:]
    ]


def _read_xlsx(path: Path) -> list[dict[str, object]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required to read .xlsx files") from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    return _rows_from_matrix([list(row) for row in ws.iter_rows(values_only=True)])


def _read_xls(path: Path) -> list[dict[str, object]]:
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("xlrd is required to read .xls files") from exc
    book = xlrd.open_workbook(path)
    sheet = book.sheet_by_index(0)
    return _rows_from_matrix([sheet.row_values(i) for i in range(sheet.nrows)])


def read_jpx_rows(path: str | Path) -> list[dict[str, object]]:
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv(path)
    if suffix == ".xlsx":
        return _read_xlsx(path)
    if suffix == ".xls":
        return _read_xls(path)
    raise ValueError(f"unsupported JPX snapshot format: {suffix}")


def domestic_company_rows(rows: Iterable[dict[str, object]]) -> list[dict[str, object]]:
    out = []
    for row in rows:
        market = str(row.get("市場・商品区分", ""))
        if DOMESTIC_MARKET_MARKER not in market:
            continue
        code = str(row.get("コード", "")).strip()
        name = str(row.get("銘柄名", "")).strip()
        if not code or not name:
            continue
        out.append(row)
    return out


def build_pilot(
    path: str | Path,
    *,
    snapshot: str,
    source_url: str,
    retrieved_at: str,
    limit: int = 100,
) -> dict:
    path = Path(path)
    rows = domestic_company_rows(read_jpx_rows(path))
    rows = sorted(rows, key=lambda r: str(r.get("コード", "")))[:limit]
    source = SourceRef(
        source="JPX",
        source_key=path.name,
        snapshot=snapshot,
        url=source_url,
        retrieved_at=retrieved_at,
        adapter_version="0.3",
    )
    entities: list[EntityRecord] = [from_jpx_row(row, source) for row in rows]
    return {
        "manifest": {
            "source": "JPX",
            "snapshot": snapshot,
            "source_url": source_url,
            "retrieved_at": retrieved_at,
            "source_file": path.name,
            "source_sha256": sha256_file(path),
            "adapter_version": "0.3",
            "selection": "sorted domestic listed equities by security code",
            "limit": limit,
            "entity_count": len(entities),
        },
        "entities": [entity.to_dict() for entity in entities],
    }


def write_pilot(payload: dict, output: str | Path) -> None:
    path = Path(output)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
