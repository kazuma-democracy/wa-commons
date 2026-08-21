from __future__ import annotations

from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
import hashlib
import re

from openpyxl import load_workbook

from wa_commons.identity.benchmark import BenchmarkRecord
from wa_commons.identity.policy import final_policy_match

SOURCE_ID = "jp-mod-procurement"
SOURCE_URL = "https://www.mod.go.jp/j/budget/chotatsu/naikyoku/keiyaku/fy2026/04_buppin_k.xlsx"
SOURCE_PAGE_URL = "https://www.mod.go.jp/j/budget/chotatsu/naikyoku/keiyaku/index.html"
SNAPSHOT_VERSION = "fy2026-04-buppin-competitive"
ADAPTER_VERSION = "0.1"
CONTRACTING_AUTHORITY = "Japan Ministry of Defense / Minister's Secretariat, Accounts Division"


@dataclass(frozen=True)
class ProcurementObservation:
    observation_id: str
    subject: str
    supplier_name: str
    supplier_address: str
    corporate_number: str
    contract_date: str
    contract_amount_jpy: int | None
    planned_price_jpy: int | None
    contracting_authority: str
    source_url: str
    source_page_url: str
    source_locator: str
    retrieved_at: str
    source_sha256: str
    snapshot_version: str = SNAPSHOT_VERSION
    adapter_version: str = ADAPTER_VERSION
    identity_decision: str = "UNRESOLVED"
    entity_id: str | None = None

    def to_dict(self) -> dict:
        return asdict(self)


def sha256(path: str | Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        while chunk := fh.read(1024 * 1024):
            h.update(chunk)
    return h.hexdigest()


def _text(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\u3000", " ")).strip()


def _digits(value: object) -> str:
    return re.sub(r"\D", "", _text(value))


def _amount(value: object) -> int | None:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return int(value)
    text = _text(value)
    if not text or "非公表" in text:
        return None
    digits = re.sub(r"[^0-9]", "", text)
    return int(digits) if digits else None


def _contract_date(value: object, fiscal_year: int = 2026) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = _text(value)
    m = re.search(r"(?:(\d{4})年)?\s*(\d{1,2})月\s*(\d{1,2})日", text)
    if not m:
        return ""
    year = int(m.group(1)) if m.group(1) else fiscal_year
    return date(year, int(m.group(2)), int(m.group(3))).isoformat()


def _split_supplier(value: object) -> tuple[str, str]:
    raw = str(value or "").replace("\r", "\n")
    parts = [p.strip() for p in raw.split("\n") if p.strip()]
    if not parts:
        return "", ""
    return parts[0], " ".join(parts[1:])


def _header_map(values: list[object]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for i, value in enumerate(values):
        text = _text(value)
        if "物品役務等の名称" in text:
            mapping["subject"] = i
        elif "契約担当官等" in text:
            mapping["authority"] = i
        elif "契約を締結した日" in text:
            mapping["date"] = i
        elif "契約の相手方" in text:
            mapping["supplier"] = i
        elif text == "法人番号" or "法人番号" in text:
            mapping["corporate_number"] = i
        elif text == "予定価格" or "予定価格" in text:
            mapping["planned_price"] = i
        elif text == "契約金額" or "契約金額" in text:
            mapping["contract_amount"] = i
    return mapping


def _find_header(ws) -> tuple[int, dict[str, int]]:
    for row_no, row in enumerate(ws.iter_rows(values_only=True), start=1):
        mapping = _header_map(list(row))
        if {"subject", "date", "supplier", "corporate_number", "contract_amount"}.issubset(mapping):
            return row_no, mapping
    raise ValueError("MOD procurement workbook header not found")


def resolve_supplier(supplier_name: str, corporate_number: str) -> tuple[str, str | None]:
    """Route supplier identity through the M1.2 decision layer.

    The source-provided Japanese corporate number is treated as a strong ID. A
    missing/malformed number remains unresolved; fuzzy name-only matching is not
    allowed to establish a consequential identity link.
    """
    number = _digits(corporate_number)
    if len(number) != 13:
        return "UNRESOLVED", None
    source = BenchmarkRecord(name=supplier_name, corporate_number=number, jurisdiction="JP")
    canonical = BenchmarkRecord(name=supplier_name, corporate_number=number, jurisdiction="JP")
    result = final_policy_match(source, canonical)
    if result.decision != "AUTO_LINK":
        return result.decision, None
    return result.decision, f"jp:corporate-number:{number}"


def parse_workbook(path: str | Path, *, retrieved_at: str) -> list[ProcurementObservation]:
    path = Path(path)
    digest = sha256(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    observations: list[ProcurementObservation] = []
    for ws in wb.worksheets:
        try:
            header_row, columns = _find_header(ws)
        except ValueError:
            continue
        for row_no, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            values = list(row)
            subject = _text(values[columns["subject"]] if columns["subject"] < len(values) else "")
            if not subject or subject.startswith("公益法人") or subject.startswith("物品役務等の名称"):
                continue
            supplier_cell = values[columns["supplier"]] if columns["supplier"] < len(values) else ""
            supplier_name, supplier_address = _split_supplier(supplier_cell)
            corporate_number = _digits(values[columns["corporate_number"]] if columns["corporate_number"] < len(values) else "")
            contract_date = _contract_date(values[columns["date"]] if columns["date"] < len(values) else "")
            if not supplier_name or not contract_date:
                continue
            authority = _text(values[columns.get("authority", -1)]) if columns.get("authority", -1) >= 0 else CONTRACTING_AUTHORITY
            decision, entity_id = resolve_supplier(supplier_name, corporate_number)
            observations.append(
                ProcurementObservation(
                    observation_id=f"wc:obs:mod-fy2026-04:{ws.title}:{row_no}",
                    subject=subject,
                    supplier_name=supplier_name,
                    supplier_address=supplier_address,
                    corporate_number=corporate_number,
                    contract_date=contract_date,
                    contract_amount_jpy=_amount(values[columns["contract_amount"]]),
                    planned_price_jpy=_amount(values[columns["planned_price"]]) if "planned_price" in columns else None,
                    contracting_authority=authority or CONTRACTING_AUTHORITY,
                    source_url=SOURCE_URL,
                    source_page_url=SOURCE_PAGE_URL,
                    source_locator=f"sheet={ws.title};row={row_no}",
                    retrieved_at=retrieved_at,
                    source_sha256=digest,
                    identity_decision=decision,
                    entity_id=entity_id,
                )
            )
    return observations


def observation_to_claim(observation: ProcurementObservation) -> dict | None:
    """Transform only resolved observations into narrow EvidenceClaim-shaped data.

    This deliberately asserts only receipt of a MOD contract. The exact subject
    is preserved in the value object. No weapons_activity or policy decision is
    inferred here.
    """
    if observation.identity_decision != "AUTO_LINK" or not observation.entity_id:
        return None
    # Source sheet names may contain Japanese characters. Canonical claim/evidence
    # IDs are ASCII-only, so derive a stable key from the full observation ID while
    # keeping the human-readable workbook locator separately in evidence.locator.
    stable_key = hashlib.sha256(observation.observation_id.encode("utf-8")).hexdigest()[:16]
    value = {
        "contracting_authority": observation.contracting_authority,
        "contract_subject": observation.subject,
        "contract_amount_jpy": observation.contract_amount_jpy,
        "supplier_name_as_published": observation.supplier_name,
        "corporate_number": observation.corporate_number,
    }
    return {
        "schema_version": "0.1",
        "claim_id": f"wc:claim:mod-fy2026-04:{stable_key}",
        "subject": {
            "entity_id": observation.entity_id,
            "entity_type": "company",
            "canonical_name": observation.supplier_name,
            "jurisdiction": "JP",
            "identifiers": [{"scheme": "corporate_number", "value": observation.corporate_number, "issuer": "National Tax Agency, Japan"}],
            "entity_resolution": {
                "method": "deterministic_identifier",
                "confidence": 1.0,
                "review_status": "not_required",
                "match_evidence": ["corporate_number", "wa-conservative-v0.2"],
            },
        },
        "claim": {
            "category": "military_contract",
            "predicate": "received_contract_from_japan_ministry_of_defense",
            "value": value,
            "currency": "JPY" if observation.contract_amount_jpy is not None else None,
            "effective_from": observation.contract_date,
            "effective_to": observation.contract_date,
        },
        "evidence": [{
            "evidence_id": f"wc:evidence:mod-fy2026-04:{stable_key}",
            "source_id": SOURCE_ID,
            "source_url": observation.source_url,
            "publisher": "Japan Ministry of Defense",
            "source_type": "official_contract",
            "publication_date": None,
            "evidence_date": observation.contract_date,
            "retrieved_at": observation.retrieved_at,
            "support": "supports",
            "locator": observation.source_locator,
            "content_sha256": observation.source_sha256,
            "license_status": "review_required",
            "notes": "Narrow contract fact only; contract subject is preserved verbatim and is not classified as weapons activity by this adapter.",
        }],
        "adjudication": {
            "status": "confirmed",
            "confidence": 1.0,
            "reasoning_summary": "Official MOD contract record plus source-published Japanese corporate number; no contract-to-weapons inference performed.",
            "method": {"type": "deterministic_rule", "version": ADAPTER_VERSION, "model": None},
            "rule_set_version": "wa-conservative-v0.2+mod-procurement-v0.1",
            "decided_at": observation.retrieved_at,
            "reviewer": None,
        },
        "policy_context": None,
        "correction_history": [],
        "provenance": {
            "created_at": observation.retrieved_at,
            "updated_at": observation.retrieved_at,
            "created_by": "wa-commons:mod-procurement-adapter",
            "dataset_version": observation.snapshot_version,
        },
    }
